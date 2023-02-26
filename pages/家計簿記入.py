import streamlit as st
import time
import datetime
import openpyxl as px
import pandas as pd

st.set_page_config(layout="wide")

inputDate=st.date_input(label='When')
inputText = st.text_input(label='Please input text', value='スーパーアルプス')
inputValue = st.text_input(label='Please input value', value='100')
inputValue=int(inputValue)
listType=["交通費","ガソリン","食費","雑貨","その他"]
inputType=st.selectbox("分類",listType)
inputHuman1=st.checkbox(label="美織")
inputHuman2=st.checkbox(label="洋太")
submit_btn=st.button("送信")
listDate1=[]
listDate2=[]
wb=px.load_workbook("./sasaki.xlsx")
now = datetime.date.today()



for i in range(7,13):
    a=datetime.date(2022,i,16)
    listDate1.append(a)

for i in range(1,3):
    a=datetime.date(2023,i,16)
    listDate1.append(a)


for i in range(8,13):
    a=datetime.date(2022,i,15)
    listDate2.append(a)

for i in range(1,4):
    a=datetime.date(2023,i,15)
    listDate2.append(a)

if submit_btn:
    if inputHuman1:


        if inputDate>listDate1[0] and inputDate<listDate2[0]:
            t=21
            ws=wb.worksheets[1]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[0])
            st.write(listDate2[0])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=1)
            


        elif inputDate>listDate1[1] and inputDate<listDate2[1]:
            t=21
            ws=wb.worksheets[2]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[1])
            st.write(listDate2[1])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=2)
           
            
            
        elif inputDate>listDate1[2] and inputDate<listDate2[2]:
            t=21
            ws=wb.worksheets[3]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[2])
            st.write(listDate2[2])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=3)
            
           
        
        elif inputDate>listDate1[3] and inputDate<listDate2[3]:
            t=21
            ws=wb.worksheets[4]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[3])
            st.write(listDate2[3])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=4)
          
            
        
        elif inputDate>listDate1[4] and inputDate<listDate2[4]:
            t=21
            ws=wb.worksheets[5]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[4])
            st.write(listDate2[4])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=5)
            
            
           
        
        elif inputDate>listDate1[5] and inputDate<listDate2[5]:
            t=21
            ws=wb.worksheets[6]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[5])
            st.write(listDate2[5])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=6)
         
            
            
        elif inputDate>listDate1[6] and inputDate<listDate2[6]:
            t=21
            ws=wb.worksheets[7]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[6])
            st.write(listDate2[6])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=7)
            
            
        
        elif inputDate>listDate1[7] and inputDate<listDate2[7]:
            t=21
            ws=wb.worksheets[8]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[7])
            st.write(listDate2[7])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=8)
            
            
        
        elif inputDate>listDate1[8] and inputDate<listDate2[8]:
            t=21
            ws=wb.worksheets[9]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[8])
            st.write(listDate2[8])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=9)
            
        
        df3["↓↓共用カード払い"]=df3["↓↓共用カード払い"].astype(str)
        df3["↓↓美織払い"]=df3["↓↓美織払い"].astype(str)
        df3["↓↓洋太払い"]=df3["↓↓洋太払い"].astype(str)
        
        st.dataframe(df3,width=1500)


    if inputHuman2:


        if inputDate>listDate1[0] and inputDate<listDate2[0]:
            t=21
            ws=wb.worksheets[1]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[0])
            st.write(listDate2[0])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=1)
            


        elif inputDate>listDate1[1] and inputDate<listDate2[1]:
            t=21
            ws=wb.worksheets[2]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[1])
            st.write(listDate2[1])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=2)
           
            
            
        elif inputDate>listDate1[2] and inputDate<listDate2[2]:
            t=21
            ws=wb.worksheets[3]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[2])
            st.write(listDate2[2])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=3)
            
           
        
        elif inputDate>listDate1[3] and inputDate<listDate2[3]:
            t=21
            ws=wb.worksheets[4]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[3])
            st.write(listDate2[3])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=4)
          
            
        
        elif inputDate>listDate1[4] and inputDate<listDate2[4]:
            t=21
            ws=wb.worksheets[5]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[4])
            st.write(listDate2[4])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=5)
            
            
           
        
        elif inputDate>listDate1[5] and inputDate<listDate2[5]:
            t=21
            ws=wb.worksheets[6]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[5])
            st.write(listDate2[5])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=6)
         
            
            
        elif inputDate>listDate1[6] and inputDate<listDate2[6]:
            t=21
            ws=wb.worksheets[7]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[6])
            st.write(listDate2[6])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=7)
            
            
        
        elif inputDate>listDate1[7] and inputDate<listDate2[7]:
            t=21
            ws=wb.worksheets[8]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[7])
            st.write(listDate2[7])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=8)
            
            
        
        elif inputDate>listDate1[8] and inputDate<listDate2[8]:
            t=21
            ws=wb.worksheets[9]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[8])
            st.write(listDate2[8])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki.xlsx")
            df3=pd.read_excel("./sasaki.xlsx",skiprows=20,sheet_name=9)
            
        
            
        df3["↓↓共用カード払い"]=df3["↓↓共用カード払い"].astype(str)
        df3["↓↓美織払い"]=df3["↓↓美織払い"].astype(str)
        df3["↓↓洋太払い"]=df3["↓↓洋太払い"].astype(str)
        st.dataframe(df3,width=1500)











    



